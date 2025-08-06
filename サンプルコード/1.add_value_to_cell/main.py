# Excel操作用のライブラリをインポート
import openpyxl
import sys
import os
import re

def add_value_to_cell(file_path, sheet_name, cell, value):
    """
    指定されたExcelセルの現在の値に、新しい数値を足し算して上書き保存する関数
    
    引数:
        file_path (str): 操作するExcelファイルのパス (例: "docs/test1.xlsx")
        sheet_name (str): 操作するシート名 (例: "アリスト")
        cell (str): セル番地 (例: "E2")
        value (float): 加算したい数値 (例: 100)
    """
    
    # 1. 指定されたファイルを開く（ファイルがなければ新規作成）
    if os.path.exists(file_path):
        # ファイルが存在する場合は読み込み
        workbook = openpyxl.load_workbook(file_path)
        print(f"既存のファイル '{file_path}' を開きました。")
    else:
        # ファイルが存在しない場合は新規作成
        workbook = openpyxl.Workbook()
        print(f"新しいファイル '{file_path}' を作成しました。")
    
    # 2. 指定されたシートを選択する（シートがなければ新規作成）
    if sheet_name in workbook.sheetnames:
        # シートが存在する場合は選択
        worksheet = workbook[sheet_name]
        print(f"既存のシート '{sheet_name}' を選択しました。")
    else:
        # シートが存在しない場合は新規作成
        worksheet = workbook.create_sheet(sheet_name)
        print(f"新しいシート '{sheet_name}' を作成しました。")
    
    # 3. 指定されたセルの値を取得する
    current_value = worksheet[cell].value
    print(f"セル {cell} の現在の値: {current_value}")
    
    # 4. もしセルが空(None)なら、値を0として扱う
    if current_value is None:
        current_value = 0
        print(f"セル {cell} は空だったので、0として扱います。")
    
    # 5. 現在の値と引数valueを足し算する
    new_value = current_value + value
    print(f"計算結果: {current_value} + {value} = {new_value}")
    
    # 6. 計算結果を同じセルに書き戻す
    worksheet[cell] = new_value
    print(f"セル {cell} に新しい値 {new_value} を設定しました。")
    
    # 7. ファイルを保存する
    # フォルダが存在しない場合は作成
    folder_path = os.path.dirname(file_path)
    if folder_path and not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"フォルダ '{folder_path}' を作成しました。")
    
    workbook.save(file_path)
    print(f"ファイル '{file_path}' を保存しました。")


# ===============================================
# ここから下は、メイン処理（実際のデータ処理）
# ===============================================

if __name__ == "__main__":
    print("Excel加算プログラムを開始します...")
    
    try:
        # 1. ファイルの確認とシート名一覧の表示
        print("\n--- Excelファイル確認 ---")
        try:
            input_workbook = openpyxl.load_workbook("docs/test1.xlsx")
            print(f"利用可能なシート: {input_workbook.sheetnames}")
        except FileNotFoundError:
            print("エラー: docs/test1.xlsx ファイルが見つかりません。")
            print("ファイルが存在することを確認してください。")
            sys.exit(1)
        
        # 2. 任意のシート名の対話型入力
        print("\n--- 任意のシート選択 ---")
        while True:
            sheet_name = input("処理対象のシート名を入力してください: ")
            
            if sheet_name.strip() == "":
                print("エラー: シート名を入力してください。")
                continue
            
            if sheet_name in input_workbook.sheetnames:
                print(f"✓ シート '{sheet_name}' を選択しました。")
                break
            else:
                print(f"エラー: 指定されたシート '{sheet_name}' が見つかりません。")
                print(f"利用可能なシート: {input_workbook.sheetnames}")
                retry = input("もう一度入力しますか？ (y/n): ")
                if retry.lower() != 'y':
                    print("プログラムを終了します。")
                    sys.exit(0)
        
        # 3. セル番地の対話型入力
        print("\n--- セル番地選択 ---")
        while True:
            cell_address = input("加算対象のセル番地を入力してください (例: A1, B2, C3): ")
            
            if cell_address.strip() == "":
                print("エラー: セル番地を入力してください。")
                continue
            
            # セル番地の簡単なバリデーション（英字+数字の形式かチェック）
            if re.match(r'^[A-Z]+[0-9]+$', cell_address.upper()):
                cell_address = cell_address.upper()  # 大文字に統一
                print(f"✓ セル番地 '{cell_address}' を選択しました。")
                break
            else:
                print("エラー: 正しいセル番地の形式で入力してください。")
                print("例: A1, B2, C3, AA10 など")
                retry = input("もう一度入力しますか？ (y/n): ")
                if retry.lower() != 'y':
                    print("プログラムを終了します。")
                    sys.exit(0)
        
        # 4. 標準入力からの数値取得
        print("\n--- 加算する数値の入力 ---")
        while True:
            try:
                user_input = input(f"セル '{cell_address}' に加算する数値を入力してください: ")
                
                # 入力値を数値に変換してバリデーション
                input_value = float(user_input)
                print(f"✓ 加算する数値: {input_value}")
                break  # 正常に数値が入力されたらループを抜ける
                
            except ValueError:
                print("エラー: 数値を入力してください。")
                print("例: 100、-50、3.14 など")
                # ループを続けて再入力を求める
        
        
        # 5. 選択されたシートとセルへの加算
        print(f"\n--- {sheet_name}シートに加算中 ---")
        add_value_to_cell(
            file_path="docs/test1.xlsx",
            sheet_name=sheet_name,
            cell=cell_address,
            value=input_value
        )
        
        # 6. 完了報告
        print("\n✓ 選択したシートに数値を加算しました。")
        print(f"  - 対象シート: {sheet_name}")
        print(f"  - 対象セル: {cell_address}")
        print(f"  - 加算した値: {input_value}")
        print("処理が正常に完了しました。")
        
    except KeyboardInterrupt:
        # Ctrl+Cで中断された場合
        print("\n\nプログラムが中断されました。")
        sys.exit(0)
    except Exception as e:
        print(f"\n予期しないエラーが発生しました: {e}")
        print("プログラムを終了します。")
        sys.exit(1)