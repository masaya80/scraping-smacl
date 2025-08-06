import openpyxl
import os
from typing import Dict


def get_cell_by_product_number(file_path: str, sheet_name: str, product_number: str, column: str) -> Dict:
    """
    指定された商品番号が何行目の何列名にあるかを取得する関数
    
    Args:
        file_path (str): エクセルのファイル名（パス）
        sheet_name (str): エクセルのシート名
        product_number (str): 検索する商品番号
        column (str): 検索対象の列名(A,B,C,D等)
    
    Returns:
        Dict: 検索結果を含む辞書
            - success (bool): 検索が成功したかどうか
            - exists (bool): 商品番号がマスターに存在するかどうか
            - row (int): 見つかった行番号（1から開始）
            - column (str): 見つかった列名
            - cell_address (str): セル番地（例：A5）
            - message (str): 結果メッセージ
            - error (str): エラーメッセージ（エラー時のみ）
    """
    
    # 結果を格納する辞書を初期化
    result = {
        'success': False,
        'exists': False,
        'row': None,
        'column': None,
        'cell_address': None,
        'message': '',
        'error': ''
    }
    
    try:
        # ①マスターファイルの存在チェック
        if not os.path.exists(file_path):
            result['error'] = f"指定されたファイルが見つかりません: {file_path}"
            result['message'] = "マスターファイルが存在しません"
            return result
        
        # Excelファイルを開く
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            result['error'] = f"ファイルの読み込みに失敗しました: {str(e)}"
            result['message'] = "ファイルの読み込みエラー"
            return result
        
        # シートの存在チェック
        if sheet_name not in workbook.sheetnames:
            result['error'] = f"指定されたシートが見つかりません: {sheet_name}"
            result['message'] = f"利用可能なシート: {', '.join(workbook.sheetnames)}"
            workbook.close()
            return result
        
        # 指定されたシートを取得
        worksheet = workbook[sheet_name]
        
        # 列名を列番号に変換
        try:
            column_index = openpyxl.utils.column_index_from_string(column.upper())
        except ValueError:
            result['error'] = f"無効な列名です: {column}"
            result['message'] = "列名はA、B、C、Dなどの形式で指定してください"
            workbook.close()
            return result
        
        # 指定された列で商品番号を検索
        found = False
        max_row = worksheet.max_row
        
        print(f"検索開始: 商品番号'{product_number}'を列{column}（{column_index}列目）で検索中...")
        print(f"対象範囲: 1行目から{max_row}行目まで")
        
        for row_num in range(1, max_row + 1):
            cell_value = worksheet.cell(row=row_num, column=column_index).value
            
            # セルの値を文字列として比較（空の場合はスキップ）
            if cell_value is not None:
                cell_value_str = str(cell_value).strip()
                product_number_str = str(product_number).strip()
                
                print(f"  行{row_num}: '{cell_value_str}' と '{product_number_str}' を比較")
                
                if cell_value_str == product_number_str:
                    # 商品番号が見つかった
                    result['success'] = True
                    result['exists'] = True
                    result['row'] = row_num
                    result['column'] = column.upper()
                    result['cell_address'] = f"{column.upper()}{row_num}"
                    result['message'] = f"商品番号'{product_number}'が見つかりました"
                    found = True
                    print(f"  ✓ 見つかりました！ セル位置: {result['cell_address']}")
                    break
        
        # 商品番号が見つからなかった場合
        if not found:
            result['success'] = True  # 検索自体は成功
            result['exists'] = False
            result['message'] = f"商品番号'{product_number}'は列{column}に存在しません"
            print(f"  商品番号'{product_number}'は見つかりませんでした")
        
        workbook.close()
        
    except Exception as e:
        result['error'] = f"予期しないエラーが発生しました: {str(e)}"
        result['message'] = "処理中にエラーが発生しました"
    
    return result


def display_search_result(result: Dict) -> None:
    """検索結果を見やすい形で表示する補助関数"""
    
    print("\n" + "="*50)
    print("🔍 商品番号検索結果")
    print("="*50)
    
    if not result['success']:
        print(f"❌ エラー: {result['error']}")
        print(f"💡 {result['message']}")
        return
    
    if result['exists']:
        print("✅ 商品番号が見つかりました！")
        print("📍 位置情報:")
        print(f"   - セル番地: {result['cell_address']}")
        print(f"   - 行番号: {result['row']}行目")
        print(f"   - 列名: {result['column']}列")
        print(f"💬 {result['message']}")
    else:
        print("❌ 商品番号が見つかりませんでした")
        print(f"💬 {result['message']}")
    
    print("="*50 + "\n")


# 使用例とテスト関数
def test_function():
    """関数のテスト用サンプル"""
    
    print("🧪 関数のテストを開始します...\n")
    
    # 実際のデータを使ったテストケース
    test_cases = [
        {
            'file_path': 'docs/test3.xlsx',
            'sheet_name': '商品マスタ',
            'product_number': '704720',  # 実際に存在する魚上魚類商品コード
            'column': 'A',
            'description': '存在する商品コード（A列）の検索'
        },
        {
            'file_path': 'docs/test3.xlsx',
            'sheet_name': '商品マスタ',
            'product_number': '5733',  # 実際に存在する東市商品コード
            'column': 'B',
            'description': '存在する商品コード（B列）の検索'
        },
        {
            'file_path': 'docs/test3.xlsx',
            'sheet_name': '商品マスタ',
            'product_number': '999999',  # 存在しない商品コード
            'column': 'A',
            'description': '存在しない商品コードの検索（見つからないケース）'
        },
        {
            'file_path': 'docs/test3.xlsx',
            'sheet_name': '商品マスタ',
            'product_number': 'ホウスイ',  # 倉庫名での検索
            'column': 'C',
            'description': '倉庫名での検索'
        }
    ]
    
    for i, test_case in enumerate(test_cases, 1):
        print(f"📝 テストケース {i}: {test_case['description']}")
        print(f"   ファイル: {test_case['file_path']}")
        print(f"   シート: {test_case['sheet_name']}")
        print(f"   商品番号: {test_case['product_number']}")
        print(f"   列: {test_case['column']}")
        
        result = get_cell_by_product_number(
            test_case['file_path'],
            test_case['sheet_name'], 
            test_case['product_number'],
            test_case['column']
        )
        
        display_search_result(result)


def usage_example():
    """実際の使用例を示す関数"""
    
    print("🎯 実際の使用例 - 受注伝票での商品番号検索\n")
    
    # 例: 受注伝票に記載された商品番号「704720」を商品マスタから検索
    product_code = "704720"
    result = get_cell_by_product_number(
        file_path='docs/test3.xlsx',
        sheet_name='商品マスタ',
        product_number=product_code,
        column='A'  # 魚上魚類商品コード列で検索
    )
    
    print(f"受注伝票の商品番号「{product_code}」をマスタで検索:")
    display_search_result(result)
    
    # 検索結果に基づく処理の例
    if result['exists']:
        print("✅ マスタに存在 → 処理続行可能")
        print(f"💡 位置: {result['cell_address']} ({result['row']}行目)")
        print("🔄 次のステップ: この行の他の情報（東市商品コード等）を取得可能\n")
    else:
        print("❌ マスタに存在しない → エラー処理が必要")
        print("🔄 次のステップ: 手動確認またはエラー通知\n")


if __name__ == "__main__":
    # メイン実行時のサンプル
    print("="*60)
    print("📋 商品番号検索システム - 動作テスト")
    print("="*60 + "\n")
    
    # テスト実行
    test_function()
    
    # 使用例の実行
    usage_example()
    
    print("="*60)
    print("🏁 テスト完了！")
    print("="*60)