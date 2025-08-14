#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 処理モジュール
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
except ImportError:
    openpyxl = None

from utils.logger import Logger
from data.pdf_extractor import DeliveryDocument, DeliveryItem


@dataclass
class MasterItem:
    """マスタアイテムのデータクラス"""
    item_code: str = ""
    item_name: str = ""
    supplier: str = ""
    category: str = ""
    unit_price: float = 0.0
    delivery_destinations: List[str] = None
    warehouse: str = ""  # 倉庫名を追加
    notes: str = ""
    
    def __post_init__(self):
        if self.delivery_destinations is None:
            self.delivery_destinations = []


@dataclass
class ValidationError:
    """バリデーションエラーのデータクラス"""
    error_type: str = ""
    item_name: str = ""
    expected_value: str = ""
    actual_value: str = ""
    description: str = ""
    document_id: str = ""


class ExcelProcessor:
    """Excel 処理クラス"""
    
    def __init__(self):
        self.logger = Logger(__name__)
        self.master_data: Dict[str, MasterItem] = {}
        
        if not openpyxl:
            raise ImportError("openpyxl がインストールされていません。pip install openpyxl を実行してください。")
    
    def load_master_data(self, master_file_path: Path) -> bool:
        """マスタExcelファイルの「登録商品マスター」シートを読み込み"""
        try:
            self.logger.info(f"マスタファイル読み込み開始: {master_file_path}")
            
            if not master_file_path.exists():
                self.logger.error(f"マスタファイルが見つかりません: {master_file_path}")
                return False
            
            # 「登録商品マスター」シートを読み込み
            df = pd.read_excel(master_file_path, sheet_name='登録商品マスター')
            
            # カラム名を正規化
            df.columns = [str(col).strip() for col in df.columns]
            self.logger.info(f"マスタシートのカラム: {list(df.columns)}")
            
            # A列（角上魚類商品コード）をキーとして辞書に変換
            for _, row in df.iterrows():
                item = MasterItem()
                
                # A列の商品コード（数値型）
                item.item_code = str(int(row.iloc[0])) if pd.notna(row.iloc[0]) else ""
                
                # 他の列の情報も取得
                item.item_name = self._get_value_from_row(row, ['商品名', '品名'])
                item.supplier = "角上魚類"  # デフォルト
                item.warehouse = self._get_value_from_row(row, ['倉庫名'])  # 倉庫名を取得
                
                # その他の情報
                if '担当者' in df.columns:
                    item.notes = self._get_value_from_row(row, ['担当者'])
                
                # 商品コードをキーとして保存（A列の値）
                if item.item_code:
                    self.master_data[item.item_code] = item
                    self.logger.debug(f"マスタ登録: コード={item.item_code}, 商品名={item.item_name}, 倉庫={item.warehouse}")
            
            self.logger.info(f"マスタデータ読み込み完了: {len(self.master_data)} 件")
            return True
            
        except Exception as e:
            self.logger.error(f"マスタファイル読み込みエラー: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _get_value_from_row(self, row: pd.Series, possible_columns: List[str]) -> str:
        """行から指定された可能なカラム名のいずれかの値を取得"""
        for col in possible_columns:
            if col in row.index:
                value = str(row[col]).strip()
                if value and value != 'nan':
                    return value
        return ""
    
    def _get_numeric_value_from_row(self, row: pd.Series, possible_columns: List[str]) -> float:
        """行から数値データを取得"""
        for col in possible_columns:
            if col in row.index:
                try:
                    value = pd.to_numeric(row[col], errors='coerce')
                    if pd.notna(value):
                        return float(value)
                except:
                    continue
        return 0.0
    
    def validate_with_master(self, documents: List[DeliveryDocument], master_file_path: Path) -> Tuple[List[DeliveryDocument], List[ValidationError]]:
        """納品データをマスタと付け合わせて検証"""
        try:
            self.logger.info("マスタ付け合わせ処理開始")
            
            # マスタデータ読み込み
            if not self.master_data:
                if not self.load_master_data(master_file_path):
                    return documents, []
            
            validated_documents = []
            all_errors = []
            
            for document in documents:
                validated_items = []
                document_errors = []
                
                for item in document.items:
                    # マスタデータと照合
                    validation_result = self._validate_item_with_master(item, document.document_id)
                    
                    if validation_result["is_valid"]:
                        # マスタ情報で補完
                        enhanced_item = self._enhance_item_with_master(item, validation_result["master_item"])
                        validated_items.append(enhanced_item)
                    else:
                        # エラー情報を記録
                        document_errors.extend(validation_result["errors"])
                        # エラーがあってもアイテムは含める（エラー情報付き）
                        item.notes = f"エラー: {', '.join([err.description for err in validation_result['errors']])}"
                        validated_items.append(item)
                
                # 検証済み文書を作成
                validated_document = document
                validated_document.items = validated_items
                validated_documents.append(validated_document)
                all_errors.extend(document_errors)
            
            self.logger.info(f"マスタ付け合わせ完了: 正常 {len(validated_documents)} 件, エラー {len(all_errors)} 件")
            return validated_documents, all_errors
            
        except Exception as e:
            self.logger.error(f"マスタ付け合わせエラー: {str(e)}")
            return documents, []
    
    def _validate_item_with_master(self, item: DeliveryItem, document_id: str) -> Dict[str, Any]:
        """個別アイテムのマスタ検証（商品コードベース）"""
        result = {
            "is_valid": True,
            "errors": [],
            "master_item": None
        }
        
        # 商品コードでマスタデータから検索
        master_item = self._find_master_item_by_code(item.item_code)
        
        if not master_item:
            # 商品コードが見つからない
            error = ValidationError(
                error_type="商品未登録",
                item_name=item.item_name,
                expected_value="マスタに存在する商品コード",
                actual_value=item.item_code,
                description=f"商品コード「{item.item_code}」はマスタに登録されていません",
                document_id=document_id
            )
            result["errors"].append(error)
            result["is_valid"] = False
            return result
        
        result["master_item"] = master_item
        self.logger.info(f"マスタ突合成功: コード={item.item_code}, 商品名={master_item.item_name}")
        
        # 単価チェック（もしPDFに単価情報があれば）
        if master_item.unit_price > 0 and item.unit_price > 0:
            price_diff_ratio = abs(item.unit_price - master_item.unit_price) / master_item.unit_price
            if price_diff_ratio > 0.1:  # 10%以上の差異
                error = ValidationError(
                    error_type="単価差異",
                    item_name=master_item.item_name,
                    expected_value=str(master_item.unit_price),
                    actual_value=str(item.unit_price),
                    description=f"単価が標準価格と {price_diff_ratio:.1%} 差異があります",
                    document_id=document_id
                )
                result["errors"].append(error)
        
        return result
    
    def _find_master_item_by_code(self, item_code: str) -> Optional[MasterItem]:
        """商品コードでマスタデータから商品を検索"""
        if item_code in self.master_data:
            return self.master_data[item_code]
        return None
    
    def _find_master_item(self, item_name: str) -> Optional[MasterItem]:
        """マスタデータから商品を検索（商品名ベース・従来互換）"""
        # 商品名での検索（従来の機能を維持）
        for master_item in self.master_data.values():
            if master_item.item_name == item_name:
                return master_item
        
        # 部分一致で検索
        for master_item in self.master_data.values():
            if item_name in master_item.item_name or master_item.item_name in item_name:
                return master_item
        
        return None
    
    def _enhance_item_with_master(self, item: DeliveryItem, master_item: MasterItem) -> DeliveryItem:
        """マスタ情報でアイテムを補強"""
        enhanced_item = item
        
        # マスタから正しい商品名を設定
        if master_item.item_name:
            enhanced_item.item_name = master_item.item_name
            self.logger.info(f"商品名補完: {item.item_code} -> {master_item.item_name}")
        
        # 不足情報を補完
        if not enhanced_item.item_code and master_item.item_code:
            enhanced_item.item_code = master_item.item_code
        
        if not enhanced_item.supplier and master_item.supplier:
            enhanced_item.supplier = master_item.supplier
        
        return enhanced_item
    
    def write_errors_to_master_sheet(self, errors: List[ValidationError], master_file_path: Path) -> bool:
        """エラーデータをマスタExcelの「エラーリスト」シートに書き込み"""
        try:
            self.logger.info("エラーリストシートへの書き込み開始")
            
            if not errors:
                self.logger.info("エラーデータがありません")
                return True
            
            if not master_file_path.exists():
                self.logger.error(f"マスタファイルが見つかりません: {master_file_path}")
                return False
            
            # Excelファイルを読み込み
            wb = load_workbook(master_file_path)
            
            # エラーリストシートを取得または作成
            if "エラーリスト" in wb.sheetnames:
                ws = wb["エラーリスト"]
                # 既存データをクリア（ヘッダー以外）
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
            else:
                ws = wb.create_sheet("エラーリスト")
            
            # ヘッダー作成
            headers = ["発生日時", "エラー種別", "商品コード", "商品名", "説明", "文書ID"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            # エラーデータを書き込み
            for row, error in enumerate(errors, 2):
                ws.cell(row=row, column=1, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row, column=2, value=error.error_type)
                ws.cell(row=row, column=3, value=error.actual_value)  # 商品コード
                ws.cell(row=row, column=4, value=error.item_name)
                ws.cell(row=row, column=5, value=error.description)
                ws.cell(row=row, column=6, value=error.document_id)
            
            # 列幅調整
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                # マージされたセルに対応した列文字の取得
                for cell in column:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        break
                
                # column_letterが取得できない場合はスキップ
                if not column_letter:
                    continue
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # ファイル保存
            wb.save(master_file_path)
            wb.close()
            
            self.logger.info(f"エラーリストシート更新完了: {len(errors)}件のエラーを記載")
            return True
            
        except Exception as e:
            self.logger.error(f"エラーリストシート書き込みエラー: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def create_error_excel(self, errors: List[ValidationError]) -> Optional[Path]:
        """エラーデータをExcelファイルに出力（バックアップ用）"""
        try:
            self.logger.info("エラーExcel作成開始")
            
            if not errors:
                self.logger.info("エラーデータがありません")
                return None
            
            # 出力ディレクトリを取得
            from config.settings import Config
            config = Config()
            
            # ファイル名生成
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            error_file = config.output_dir / f"エラーデータ_{timestamp}.xlsx"
            
            # ワークブック作成
            wb = Workbook()
            ws = wb.active
            ws.title = "エラーデータ"
            
            # ヘッダー作成
            headers = ["エラー種別", "商品名", "期待値", "実際値", "説明", "文書ID", "発生日時"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # データ行作成
            for row, error in enumerate(errors, 2):
                ws.cell(row=row, column=1, value=error.error_type)
                ws.cell(row=row, column=2, value=error.item_name)
                ws.cell(row=row, column=3, value=error.expected_value)
                ws.cell(row=row, column=4, value=error.actual_value)
                ws.cell(row=row, column=5, value=error.description)
                ws.cell(row=row, column=6, value=error.document_id)
                ws.cell(row=row, column=7, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            # 列幅調整
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                # マージされたセルに対応した列文字の取得
                for cell in column:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        break
                
                # column_letterが取得できない場合はスキップ
                if not column_letter:
                    continue
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # ファイル保存
            wb.save(error_file)
            
            self.logger.info(f"エラーExcel作成完了: {error_file}")
            return error_file
            
        except Exception as e:
            self.logger.error(f"エラーExcel作成エラー: {str(e)}")
            return None
    
    def group_by_destination(self, documents: List[DeliveryDocument]) -> Dict[str, List[DeliveryDocument]]:
        """納品先ごとにデータをグループ化"""
        try:
            self.logger.info("納品先別グループ化開始")
            
            grouped = {}
            
            for document in documents:
                destination = document.delivery_destination or "不明"
                
                if destination not in grouped:
                    grouped[destination] = []
                
                grouped[destination].append(document)
            
            self.logger.info(f"グループ化完了: {len(grouped)} 納品先")
            return grouped
            
        except Exception as e:
            self.logger.error(f"グループ化エラー: {str(e)}")
            return {}
    
    def create_dispatch_table(self, destination: str, documents: List[DeliveryDocument]) -> Optional[Path]:
        """配車表を作成"""
        try:
            self.logger.info(f"配車表作成開始: {destination}")
            
            from config.settings import Config
            config = Config()
            
            # ファイル名生成
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_destination = self._sanitize_filename(destination)
            dispatch_file = config.output_dir / f"配車表_{safe_destination}_{timestamp}.xlsx"
            
            # ワークブック作成
            wb = Workbook()
            ws = wb.active
            ws.title = "配車表"
            
            # タイトル
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = f"配車表 - {destination}"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # ヘッダー
            headers = ["配送日", "商品名", "数量", "単位", "配送先", "備考", "担当者"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            
            # データ行
            row = 4
            for document in documents:
                for item in document.items:
                    ws.cell(row=row, column=1, value=document.document_date)
                    ws.cell(row=row, column=2, value=item.item_name)
                    ws.cell(row=row, column=3, value=item.quantity)
                    ws.cell(row=row, column=4, value=item.unit)
                    ws.cell(row=row, column=5, value=destination)
                    ws.cell(row=row, column=6, value=item.notes)
                    ws.cell(row=row, column=7, value="")  # 担当者は空欄
                    row += 1
            
            # 罫線とフォーマット
            self._apply_table_formatting(ws, 3, row - 1, len(headers))
            
            # 列幅調整
            self._adjust_column_widths(ws)
            
            # ファイル保存
            wb.save(dispatch_file)
            
            self.logger.info(f"配車表作成完了: {dispatch_file}")
            return dispatch_file
            
        except Exception as e:
            self.logger.error(f"配車表作成エラー: {str(e)}")
            return None
    
    def create_shipping_request(self, destination: str, documents: List[DeliveryDocument]) -> Optional[Path]:
        """出庫依頼書を作成"""
        try:
            self.logger.info(f"出庫依頼書作成開始: {destination}")
            
            from config.settings import Config
            config = Config()
            
            # ファイル名生成
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_destination = self._sanitize_filename(destination)
            shipping_file = config.output_dir / f"出庫依頼_{safe_destination}_{timestamp}.xlsx"
            
            # ワークブック作成
            wb = Workbook()
            ws = wb.active
            ws.title = "出庫依頼"
            
            # タイトル
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = f"出庫依頼書 - {destination}"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # 基本情報
            ws.cell(row=2, column=1, value="依頼日:")
            ws.cell(row=2, column=2, value=datetime.now().strftime('%Y-%m-%d'))
            ws.cell(row=2, column=4, value="納品先:")
            ws.cell(row=2, column=5, value=destination)
            
            # ヘッダー
            headers = ["商品コード", "商品名", "数量", "単位", "単価", "金額"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            # データ行と合計計算
            row = 5
            total_amount = 0
            
            for document in documents:
                for item in document.items:
                    ws.cell(row=row, column=1, value=item.item_code)
                    ws.cell(row=row, column=2, value=item.item_name)
                    ws.cell(row=row, column=3, value=item.quantity)
                    ws.cell(row=row, column=4, value=item.unit)
                    ws.cell(row=row, column=5, value=item.unit_price)
                    ws.cell(row=row, column=6, value=item.total_price)
                    total_amount += item.total_price
                    row += 1
            
            # 合計行
            ws.cell(row=row, column=5, value="合計:")
            ws.cell(row=row, column=5).font = Font(bold=True)
            ws.cell(row=row, column=6, value=total_amount)
            ws.cell(row=row, column=6).font = Font(bold=True)
            
            # 罫線とフォーマット
            self._apply_table_formatting(ws, 4, row, len(headers))
            
            # 列幅調整
            self._adjust_column_widths(ws)
            
            # ファイル保存
            wb.save(shipping_file)
            
            self.logger.info(f"出庫依頼書作成完了: {shipping_file}")
            return shipping_file
            
        except Exception as e:
            self.logger.error(f"出庫依頼書作成エラー: {str(e)}")
            return None
    
    def _sanitize_filename(self, filename: str) -> str:
        """ファイル名に使用できない文字を除去"""
        import re
        return re.sub(r'[<>:"/\\|?*]', '_', filename)
    
    def _apply_table_formatting(self, ws, start_row: int, end_row: int, num_cols: int):
        """テーブルに罫線とフォーマットを適用"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    def _adjust_column_widths(self, ws):
        """列幅を自動調整"""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # マージされたセルに対応した列文字の取得
            for cell in column:
                # マージされたセルでない最初のセルから列文字を取得
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            # column_letterが取得できない場合はスキップ
            if not column_letter:
                continue
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def process_warehouse_orders(self, validated_data: List[DeliveryDocument], master_file_path: Path) -> bool:
        """倉庫別注文処理 - 要求された仕様に従って既存Excelに数量を挿入"""
        try:
            self.logger.info("倉庫別注文処理開始")
            
            # マスタファイルを開く
            if not master_file_path.exists():
                self.logger.error(f"マスタファイルが見つかりません: {master_file_path}")
                return False
            
            wb = load_workbook(master_file_path)
            
            # 登録商品マスタをロード
            if not self.load_master_data(master_file_path):
                self.logger.error("登録商品マスタの読み込みに失敗しました")
                return False
            
            # すべての注文アイテムを処理
            for document in validated_data:
                for item in document.items:
                    # 1. 登録商品マスタと突合
                    if item.item_code not in self.master_data:
                        self.logger.warning(f"商品コード {item.item_code} がマスタに見つかりません")
                        continue
                    
                    # 2. C列の倉庫名を確認（ここでは商品のdelivery_destinationを使用）
                    warehouse = self._determine_warehouse(item, document)
                    
                    if warehouse == "ホウスイ":
                        self._process_housui_order(wb, item)
                    elif warehouse == "アリスト":
                        self._process_arist_order(wb, item)
                    else:
                        self.logger.warning(f"未知の倉庫名: {warehouse}")
            
            # outputディレクトリに結果ファイルを保存
            output_file = self._create_output_filename(master_file_path)
            wb.save(output_file)
            self.logger.info(f"倉庫別注文処理完了: {output_file}")
            
            # PDF出力を実行
            self._export_target_sheets_to_pdf(output_file)
            
            return True
            
        except Exception as e:
            self.logger.error(f"倉庫別注文処理エラー: {str(e)}")
            return False
    
    def _determine_warehouse(self, item: DeliveryItem, document: DeliveryDocument) -> str:
        """倉庫名を判定"""
        # DeliveryItemの倉庫名フィールドを最優先で使用
        if item.warehouse:
            warehouse = item.warehouse.strip()
            # 倉庫名の正規化
            if "ホウスイ" in warehouse or "豊水" in warehouse or "housui" in warehouse.lower():
                return "ホウスイ"
            elif "アリスト" in warehouse or "arist" in warehouse.lower():
                return "アリスト"
            else:
                # 数値コードの場合は、マスタExcelの倉庫名列を確認
                warehouse_from_master = self._get_warehouse_from_master(item.item_code)
                if warehouse_from_master:
                    return warehouse_from_master
                
                # その他の倉庫名の場合は、デフォルトとしてホウスイを使用
                self.logger.warning(f"未知の倉庫名: {warehouse}、デフォルトでホウスイを使用")
                return "ホウスイ"
        
        # 倉庫名が設定されていない場合は、マスタから取得を試行
        warehouse_from_master = self._get_warehouse_from_master(item.item_code)
        if warehouse_from_master:
            return warehouse_from_master
        
        # 納品先ベースで判定（フォールバック）
        if "ホウスイ" in (document.delivery_destination or ""):
            return "ホウスイ"
        elif "アリスト" in (document.delivery_destination or ""):
            return "アリスト"
        else:
            # デフォルトはホウスイとする
            self.logger.info("倉庫名が判定できません。デフォルトでホウスイを使用")
            return "ホウスイ"
    
    def _get_warehouse_from_master(self, item_code: str) -> Optional[str]:
        """マスタから商品の倉庫名を取得"""
        try:
            if item_code in self.master_data:
                warehouse = self.master_data[item_code].warehouse
                if warehouse:
                    # 倉庫名の正規化
                    if "ホウスイ" in warehouse or "豊水" in warehouse:
                        return "ホウスイ"
                    elif "アリスト" in warehouse:
                        return "アリスト"
                    else:
                        self.logger.debug(f"マスタの倉庫名を正規化: {warehouse}")
                        return "ホウスイ"  # デフォルト
        except Exception as e:
            self.logger.debug(f"マスタから倉庫名取得エラー: {str(e)}")
        
        return None
    
    def _process_housui_order(self, wb, item: DeliveryItem):
        """ホウスイの場合の処理"""
        try:
            # 3. ホウスイ川島出庫依頼書シートを開く
            if "ホウスイ川島出庫依頼書" not in wb.sheetnames:
                self.logger.warning("ホウスイ川島出庫依頼書シートが見つかりません")
                return
            
            housui_sheet = wb["ホウスイ川島出庫依頼書"]
            
            # 4. A列をみて商品コードが合致する行を取得
            housui_row = self._find_row_by_product_code(housui_sheet, "A", item.item_code)
            if housui_row:
                # 5. 数量をAB列のみに挿入（A列は元のまま保持）
                self._safe_cell_insert(housui_sheet, housui_row, 28, item.quantity)  # AB列のみ
                self.logger.info(f"ホウスイ川島出庫依頼書のAB列に数量挿入: 行{housui_row}, 数量{item.quantity}")
            
            # 6. アリスト鶴ヶ島 (LT1)を開く
            if "アリスト鶴ヶ島 (LT1)" not in wb.sheetnames:
                self.logger.warning("アリスト鶴ヶ島 (LT1)シートが見つかりません")
                return
            
            arist_sheet = wb["アリスト鶴ヶ島 (LT1)"]
            
            # 7. O列をみて商品コードが合致する行を取得
            arist_row = self._find_row_by_product_code(arist_sheet, "O", item.item_code)
            if arist_row:
                # 8. 数量を該当する行のH列に挿入
                self._safe_cell_insert(arist_sheet, arist_row, 8, item.quantity)  # H列
                self.logger.info(f"アリスト鶴ヶ島 (LT1)のH列に数量挿入: 行{arist_row}, 数量{item.quantity}")
            
        except Exception as e:
            self.logger.error(f"ホウスイ注文処理エラー: {str(e)}")
    
    def _process_arist_order(self, wb, item: DeliveryItem):
        """アリストの場合の処理"""
        try:
            # 3. アリスト鶴ヶ島 (LT1)を開く
            if "アリスト鶴ヶ島 (LT1)" not in wb.sheetnames:
                self.logger.warning("アリスト鶴ヶ島 (LT1)シートが見つかりません")
                return
            
            arist_sheet = wb["アリスト鶴ヶ島 (LT1)"]
            
            # 4. O列をみて商品コードが合致する行を取得
            arist_row = self._find_row_by_product_code(arist_sheet, "O", item.item_code)
            if arist_row:
                # 5. 数量を該当する行のG列に挿入
                self._safe_cell_insert(arist_sheet, arist_row, 7, item.quantity)  # G列
                self.logger.info(f"アリスト鶴ヶ島 (LT1)のG列に数量挿入: 行{arist_row}, 数量{item.quantity}")
            
        except Exception as e:
            self.logger.error(f"アリスト注文処理エラー: {str(e)}")
    
    def _find_row_by_product_code(self, worksheet, column_letter: str, product_code: str) -> Optional[int]:
        """指定された列で商品コードが合致する行を検索"""
        try:
            # 列をA=1, B=2, ... に変換
            column_index = ord(column_letter.upper()) - ord('A') + 1
            
            # 最大行数を取得
            max_row = worksheet.max_row
            
            for row in range(1, max_row + 1):
                cell_value = worksheet.cell(row=row, column=column_index).value
                if cell_value is not None:
                    # 数値型の場合は文字列に変換して比較
                    cell_str = str(int(cell_value)) if isinstance(cell_value, (int, float)) else str(cell_value)
                    if cell_str == str(product_code):
                        return row
            
            self.logger.warning(f"商品コード {product_code} が{column_letter}列に見つかりませんでした")
            return None
            
        except Exception as e:
            self.logger.error(f"商品コード検索エラー: {str(e)}")
            return None
    
    def _safe_cell_insert(self, worksheet, row: int, column: int, value):
        """マージされたセルに安全に値を挿入（マージ構造を保持）"""
        try:
            cell = worksheet.cell(row=row, column=column)
            cell_coordinate = cell.coordinate
            
            # マージされたセルかチェック
            merged_cell_found = False
            for merged_range in worksheet.merged_cells.ranges:
                if cell_coordinate in merged_range:
                    # マージされたセルの場合、左上のセル（トップレフト）に値を設定
                    top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                    top_left_cell.value = value
                    merged_cell_found = True
                    self.logger.debug(f"マージセル対応: {cell_coordinate} -> 左上セル {top_left_cell.coordinate} に値={value} を設定")
                    break
            
            # 通常のセル（マージされていない）の場合
            if not merged_cell_found:
                cell.value = value
                self.logger.debug(f"通常セル: {cell_coordinate} に値={value} を設定")
            
        except Exception as e:
            self.logger.warning(f"セル挿入エラー (行{row}, 列{column}): {str(e)}")
            # フォールバック: 強制的に値を設定
            try:
                # 直接セルアクセスで値を設定
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(column)
                worksheet[f"{col_letter}{row}"] = value
                self.logger.debug(f"フォールバック成功: {col_letter}{row} = {value}")
            except Exception as e2:
                self.logger.error(f"セル挿入完全失敗 (行{row}, 列{column}): {str(e2)}")
    
    def _is_cells_merged(self, worksheet, row1: int, col1: int, row2: int, col2: int) -> bool:
        """指定された2つのセルが同じマージ範囲に含まれているかチェック"""
        try:
            cell1_coordinate = worksheet.cell(row1, col1).coordinate
            cell2_coordinate = worksheet.cell(row2, col2).coordinate
            
            for merged_range in worksheet.merged_cells.ranges:
                if cell1_coordinate in merged_range and cell2_coordinate in merged_range:
                    return True
            
            return False
            
        except Exception as e:
            self.logger.debug(f"マージセルチェックエラー: {str(e)}")
            return False
    
    def _create_output_filename(self, master_file_path: Path) -> Path:
        """outputディレクトリにファイル名を生成"""
        try:
            from config.settings import Config
            config = Config()
            
            # タイムスタンプを作成
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # ベースファイル名を作成
            base_name = master_file_path.stem
            output_name = f"{base_name}_updated_{timestamp}.xlsx"
            
            output_file = config.output_dir / output_name
            
            self.logger.info(f"出力ファイル名生成: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"出力ファイル名生成エラー: {str(e)}")
            # フォールバック: 元のファイルと同じディレクトリに保存
            return master_file_path.with_suffix('.output.xlsx')
    
    def export_sheets_to_pdf(self, excel_file_path: Path, sheet_names: list) -> list:
        """指定されたシートをPDFとして出力"""
        pdf_files = []
        
        try:
            from config.settings import Config
            config = Config()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            for sheet_name in sheet_names:
                try:
                    # PDFファイル名を生成
                    safe_sheet_name = self._sanitize_filename(sheet_name)
                    pdf_filename = f"{safe_sheet_name}_{timestamp}.pdf"
                    pdf_path = config.output_dir / pdf_filename
                    
                    # xlwingsを使用してExcelファイルを開きPDF出力
                    success = self._export_sheet_to_pdf_xlwings(excel_file_path, sheet_name, pdf_path)
                    
                    if success:
                        pdf_files.append(pdf_path)
                        self.logger.info(f"PDF出力成功: {sheet_name} -> {pdf_path}")
                    else:
                        self.logger.warning(f"PDF出力失敗: {sheet_name}")
                        
                except Exception as e:
                    self.logger.error(f"シート '{sheet_name}' のPDF出力エラー: {str(e)}")
            
            return pdf_files
            
        except Exception as e:
            self.logger.error(f"PDF出力処理エラー: {str(e)}")
            return []
    
    def _export_sheet_to_pdf_xlwings(self, excel_file_path: Path, sheet_name: str, pdf_path: Path) -> bool:
        """xlwingsを使用してシートをPDFに出力（設定に応じて有効/無効）"""
        try:
            # 設定を確認してExcelアプリケーションの使用を判定
            from config.settings import Config
            config = Config()
            
            # macOSなど、Excelアプリの使用が無効になっている場合は代替方法を使用
            if not config.enable_excel_app:
                self.logger.info("Excel アプリケーションの使用が無効化されています。代替方法を使用します")
                return self._export_sheet_to_pdf_alternative(excel_file_path, sheet_name, pdf_path)
            
            import xlwings as xw
            
            # Excelアプリケーションを起動（非表示）
            app = xw.App(visible=False)
            
            try:
                # ワークブックを開く
                wb = app.books.open(str(excel_file_path))
                
                # 指定されたシートを取得
                if sheet_name in [ws.name for ws in wb.sheets]:
                    ws = wb.sheets[sheet_name]
                    
                    # PDFとして出力
                    ws.to_pdf(str(pdf_path))
                    
                    self.logger.info(f"xlwingsでPDF出力: {sheet_name} -> {pdf_path}")
                    return True
                else:
                    self.logger.warning(f"シート '{sheet_name}' が見つかりません")
                    return False
                    
            finally:
                # ワークブックとアプリケーションを閉じる
                wb.close()
                app.quit()
                
        except ImportError:
            self.logger.warning("xlwingsがインストールされていません。代替方法を試行します")
            return self._export_sheet_to_pdf_alternative(excel_file_path, sheet_name, pdf_path)
        except Exception as e:
            self.logger.error(f"xlwingsでのPDF出力エラー: {str(e)}")
            return self._export_sheet_to_pdf_alternative(excel_file_path, sheet_name, pdf_path)
    
    def _export_sheet_to_pdf_alternative(self, excel_file_path: Path, sheet_name: str, pdf_path: Path) -> bool:
        """代替方法でシートをPDFに出力（openpyxl + HTML経由）"""
        try:
            # まずweasyprint経由を試行
            if self._export_sheet_to_pdf_weasyprint(excel_file_path, sheet_name, pdf_path):
                return True
            
            # weasyprint が利用できない場合は、pandasでシンプルなPDF出力を試行
            return self._export_sheet_to_pdf_pandas(excel_file_path, sheet_name, pdf_path)
            
        except Exception as e:
            self.logger.error(f"代替方法でのPDF出力エラー: {str(e)}")
            return False
    
    def _export_sheet_to_pdf_weasyprint(self, excel_file_path: Path, sheet_name: str, pdf_path: Path) -> bool:
        """weasprintを使用してPDF出力（高品質）"""
        try:
            from weasyprint import HTML, CSS
            
            # openpyxlでワークシートを読み込み（セル結合とスタイル情報を保持）
            wb = load_workbook(excel_file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                self.logger.warning(f"シート '{sheet_name}' が見つかりません")
                return False
            
            ws = wb[sheet_name]
            
            # HTMLテーブルを生成（より正確にセル結合を反映）
            html_content = self._generate_html_from_worksheet(ws, sheet_name)
            
            # CSSスタイル定義
            css_style = CSS(string="""
                @page {
                    size: A4 landscape;
                    margin: 1cm;
                }
                body {
                    font-family: Arial, "MS PGothic", sans-serif;
                    font-size: 12px;
                    margin: 0;
                }
                table {
                    border-collapse: collapse;
                    width: 100%;
                    table-layout: fixed;
                }
                th, td {
                    border: 1px solid #000;
                    padding: 4px;
                    text-align: left;
                    vertical-align: top;
                    word-wrap: break-word;
                }
                th {
                    background-color: #f0f0f0;
                    font-weight: bold;
                }
                .header-cell {
                    background-color: #e6f3ff;
                    font-weight: bold;
                }
                .title {
                    font-size: 16px;
                    font-weight: bold;
                    text-align: center;
                    margin-bottom: 10px;
                }
            """)
            
            # HTMLからPDFを生成
            HTML(string=html_content).write_pdf(str(pdf_path), stylesheets=[css_style])
            
            wb.close()
            self.logger.info(f"weasyprint でPDF出力成功: {sheet_name} -> {pdf_path}")
            return True
            
        except ImportError:
            self.logger.debug("weasyprint がインストールされていません")
            return False
        except Exception as e:
            self.logger.debug(f"weasyprint でのPDF出力エラー: {str(e)}")
            return False
    
    def _generate_html_from_worksheet(self, worksheet, sheet_name: str) -> str:
        """ワークシートからHTMLテーブルを生成（セル結合を考慮）"""
        try:
            html_rows = []
            max_row = min(worksheet.max_row, 200)  # 最大200行まで処理
            max_col = min(worksheet.max_column, 50)  # 最大50列まで処理
            
            # マージされたセル範囲を取得
            merged_ranges = list(worksheet.merged_cells.ranges)
            
            for row in range(1, max_row + 1):
                html_cells = []
                
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # マージされたセルの処理
                    skip_cell = False
                    rowspan = 1
                    colspan = 1
                    
                    for merged_range in merged_ranges:
                        if cell.coordinate in merged_range:
                            # マージされたセルの左上セル以外はスキップ
                            if cell.row != merged_range.min_row or cell.column != merged_range.min_col:
                                skip_cell = True
                                break
                            else:
                                # 左上セルの場合はrowspanとcolspanを設定
                                rowspan = merged_range.max_row - merged_range.min_row + 1
                                colspan = merged_range.max_col - merged_range.min_col + 1
                    
                    if skip_cell:
                        continue
                    
                    # セルの値を取得
                    cell_value = cell.value if cell.value is not None else ""
                    cell_value_str = str(cell_value).strip()
                    
                    # セルのクラスを決定
                    cell_class = ""
                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color.rgb:
                        cell_class = "header-cell"
                    
                    # HTMLセルを生成
                    attrs = []
                    if rowspan > 1:
                        attrs.append(f'rowspan="{rowspan}"')
                    if colspan > 1:
                        attrs.append(f'colspan="{colspan}"')
                    if cell_class:
                        attrs.append(f'class="{cell_class}"')
                    
                    attrs_str = ' ' + ' '.join(attrs) if attrs else ''
                    html_cells.append(f"<td{attrs_str}>{cell_value_str}</td>")
                
                if html_cells:
                    html_rows.append(f"<tr>{''.join(html_cells)}</tr>")
            
            # HTMLドキュメントを構成
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{sheet_name}</title>
            </head>
            <body>
                <div class="title">{sheet_name}</div>
                <table>
                    {''.join(html_rows)}
                </table>
            </body>
            </html>
            """
            
            return html_content
            
        except Exception as e:
            self.logger.error(f"HTML生成エラー: {str(e)}")
            # フォールバック: シンプルなHTML
            return f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{sheet_name}</title>
            </head>
            <body>
                <h1>{sheet_name}</h1>
                <p>データの表示でエラーが発生しました: {str(e)}</p>
            </body>
            </html>
            """
    
    def _export_sheet_to_pdf_pandas(self, excel_file_path: Path, sheet_name: str, pdf_path: Path) -> bool:
        """pandasを使用したシンプルなPDF出力（フォールバック）"""
        try:
            import pandas as pd
            
            # Excelファイルからシートを読み込み
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
            
            # HTMLに変換
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{sheet_name}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f2f2f2; }}
                    h1 {{ color: #333; }}
                </style>
            </head>
            <body>
                <h1>{sheet_name}</h1>
                {df.to_html(escape=False, index=False)}
            </body>
            </html>
            """
            
            # HTMLファイルとして一時保存し、警告メッセージを出力
            html_temp_file = pdf_path.with_suffix('.html')
            with open(html_temp_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.warning(f"PDF出力の代わりにHTMLファイルを生成しました: {html_temp_file}")
            self.logger.warning("完全なPDF出力を行うには、weasyprint をインストールしてください: pip install weasyprint")
            
            return True
            
        except Exception as e:
            self.logger.error(f"pandas でのHTML出力エラー: {str(e)}")
            return False
    
    def _export_target_sheets_to_pdf(self, excel_file_path: Path):
        """ホウスイ川島出庫依頼書とアリスト鶴ヶ島 (LT1)シートをPDFに出力"""
        try:
            target_sheets = [
                "ホウスイ川島出庫依頼書",
                "アリスト鶴ヶ島 (LT1)"
            ]
            
            self.logger.info("PDF出力開始: ホウスイ川島出庫依頼書、アリスト鶴ヶ島 (LT1)")
            
            pdf_files = self.export_sheets_to_pdf(excel_file_path, target_sheets)
            
            if pdf_files:
                self.logger.info(f"PDF出力完了: {len(pdf_files)}件のファイルが生成されました")
                for pdf_file in pdf_files:
                    self.logger.info(f"生成されたPDF: {pdf_file}")
            else:
                self.logger.warning("PDF出力に失敗しました")
                
        except Exception as e:
            self.logger.error(f"PDF出力エラー: {str(e)}") 